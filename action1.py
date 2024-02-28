import os
import re
from multiprocessing import Process
import openpyxl
from threading import Thread
import asyncio
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from seleniumwire.utils import decode as sw_decode
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.webdriver.support.ui import Select
import json
import var1
from retry import retry
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import subprocess
import pickle
import threading
from seleniumwire import webdriver
from seleniumwire.utils import decode

options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors-spki-list')
options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')
options.add_argument('log-level=3')
driver = webdriver.Chrome(var1.PATH, chrome_options=options)



# from DrissionPage import ChromiumPage
from DrissionPage import *
driver2 = ChromiumPage()
driver2.new_tab()
shopify = driver2.get_tab(2)
tiktok = driver2.get_tab(1)


from selenium.common.exceptions import InvalidSessionIdException
import unittest
import logging
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys


def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)




file_name = 'C:/Users/Admin/PycharmProjects/pythonProject/litcommerce/data_litcommerce.json'
with open(file_name, 'r', encoding='utf-8') as f:
    data = json.load(f, strict = False)


logging.basicConfig(handlers=[logging.FileHandler(filename= "C:/Users/Admin/PycharmProjects/pythonProject/litcommerce/litcommerce.log",
                                                 encoding= 'utf-8', mode='a+')],
                    format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                    datefmt="%F %A %T",
                    level=logging.INFO)


class login():
    def login_shopify_tiktok(self):
        shopify.get("https://admin.shopify.com/store/ali-shop-2503")
        time.sleep(2)
        print(shopify.tab_id)
        print(shopify.title)
        tiktok.get("https://seller-us.tiktok.com/product/manage")
        time.sleep(2)
        print(tiktok.tab_id)
        print(tiktok.title)

    def login_litcommerce(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var1.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var1.login_user).send_keys(user)
        driver.find_element(By.XPATH, var1.login_password).send_keys(password)
        driver.find_element(By.XPATH, var1.login_submit).click()
        driver.find_element(By.XPATH, var1.icon_lit).is_displayed()
        time.sleep(2)
        logging.info("Đã login thành công Litcommerce với tài khoản: " + user)
        logging.info("Đã login thành công Litcommerce với mật khẩu: " + password)
        logging.info("Title trang: "+ driver.title)
        logging.info(driver.title == "Dashboard")
        print(driver.title)
        driver.implicitly_wait(2)
        try:
            driver.find_element(By.XPATH, var1.icon_x).click()
        except:
            pass
        time.sleep(1)




class shopify_space():
    def order_listorder(self):
        print("1", shopify.title)
        shopify.ele(var1.shopify_order).click()

    def oder_addorder(self):
        time.sleep(2)
        shopify.ele(var1.shopify_addorder).click()
        a  = shopify.ele(var1.shopify_addorder).text
        print("test text", a)


class tiktok_space():
    def oder_listorder(self):
        time.sleep(2)
        print("2", tiktok.title)
        tiktok.ele(var1.tiktok_order).click()

    def oder_managerreturns(self):
        time.sleep(2)
        tiktok.ele(var1.tiktok_managereturn).click()

class litcommerce():
    pass

